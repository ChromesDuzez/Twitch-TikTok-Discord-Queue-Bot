import discord
from discord.ext import commands
from datetime import datetime, timedelta
import sqlite3
import xlsxwriter as xwriter
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment, Protection
from copy import copy
import os

def canPressButton(interaction: discord.Interaction, intended_user: discord.User, accepted_roles: list = ['TIMECARD_ADMIN_ROLE']) -> bool:
    admin_role_id = int(os.getenv('TIMECARD_ADMIN_ROLE'))
    user = interaction.user

    if user.id == intended_user.id:
        return True

    if any(role.permissions.administrator for role in user.roles):
        return True

    for role in accepted_roles:
        role_id = None
        if type(role) == type(109):
            role_id = role
        elif type(role) == type("str"):
            role_id = int(os.getenv(role))
        
        if role_id:
            if discord.utils.get(user.roles, id=role_id):
                return True

    return False

class Confirm(discord.ui.View):
    def __init__(self, user: discord.User, timeout: int = None):
        super().__init__(timeout=timeout)
        self.user = user
        self.value = None

    @discord.ui.button(label="Yes", style=discord.ButtonStyle.green)
    async def confirm(self, button: discord.ui.Button, interaction: discord.Interaction):
        if interaction.user != self.user:
            await interaction.response.send_message("This is not for you!", ephemeral=True)
            return
        self.value = True
        self.stop()
        await interaction.response.send_message("You clicked Yes!", ephemeral=True)

    @discord.ui.button(label="No", style=discord.ButtonStyle.red)
    async def cancel(self, button: discord.ui.Button, interaction: discord.Interaction):
        if interaction.user != self.user:
            await interaction.response.send_message("This is not for you!", ephemeral=True)
            return
        self.value = False
        self.stop()
        await interaction.response.send_message("You clicked No!", ephemeral=True)

class ApprovePunch(discord.ui.View):
    def __init__(self, punch: int, message: discord.message, bot: discord.bot, db:str):
        print(f"Created <ApprovePunch: Object> with values: [{punch}, {message}, {db}]")
        super().__init__(timeout=None)
        self.bot = bot
        self.db = db
        self.punch = punch
        self.message = message

        # Dynamically add buttons based on the booleans
        conn = sqlite3.connect(self.db)
        cursor = conn.cursor()
        cursor.execute(f"SELECT punchInApproval, punchOutApproval FROM punch_clock WHERE id = {punch}")
        result = cursor.fetchone()
        #print(result)#debugging
        self.punchInApproval = result[0]
        self.punchOutApproval = result[1]
        conn.commit()
        conn.close()
        if not result[0]:
            self.add_item(self.ConfirmButton("clock-in"))
            self.add_item(self.EditPunch("clock-in"))
        if not result[1]:
            self.add_item(self.ConfirmButton("clock-out"))
            self.add_item(self.EditPunch("clock-out"))
    
    class ConfirmButton(discord.ui.Button):
        def __init__(self, label: str):
            self.CustomLabel = label
            color = discord.ButtonStyle.green
            if label == "clock-out":
                color = discord.ButtonStyle.red
            super().__init__(label=f"Approve {label}", style=color)

        async def callback(self, interaction: discord.Interaction):
            view: ApprovePunch = self.view
            role = None
            if os.getenv('TIMECARD_ADMIN_ROLES'):
                role = discord.utils.get(interaction.user.roles, id=int(os.getenv('TIMECARD_ADMIN_ROLE')))
            if (not role) and (not interaction.user.guild_permissions.administrator):
                await interaction.response.send_message("This is not for you!", ephemeral=True)
                return
            if self.CustomLabel == "clock-in":
                view.punchInApproval = True
            else:
                view.punchOutApproval = True
            conn = sqlite3.connect(view.db)
            cursor = conn.cursor()
            cursor.execute(f'UPDATE punch_clock SET punchInApproval = {view.punchInApproval}, punchOutApproval = {view.punchOutApproval} WHERE id = {view.punch}')
            messageContent = view.message.content
            approvaltxt = f"**You approved this {self.CustomLabel} attempt.**"
            #print(messageContent) #debugging
            #print(f"Degugging Message: {messageContent}") #debugging
            if messageContent[-37::] == "Do you approve of this login attempt?":
                messageContent = messageContent[:-37:] + f"{approvaltxt}"
            else:
                messageContent = messageContent + f"\n{approvaltxt}"
            #print(f"Degugging New Message: {messageContent}") #debugging
            if view.punchInApproval and view.punchOutApproval:
                await interaction.message.edit(content=messageContent, view=None)
                cursor.execute(f'UPDATE punch_clock SET checkChannelId = NULL, checkMessageId = NULL WHERE id = {view.punch}')
                conn.commit()
                conn.close()
            else:
                conn.commit()
                conn.close()
                await interaction.message.edit(content=messageContent, view=ApprovePunch(punch=view.punch, message=view.message, bot=view.bot, db=view.db))
            await interaction.response.send_message(content=approvaltxt, ephemeral=True)
            return

    class EditPunch(discord.ui.Button):
        def __init__(self, label: str):
            self.CustomLabel = label
            color = discord.ButtonStyle.secondary #colors: [primary, secondary, green, red, link]
            # if label == "clock-out":
            #     color = discord.ButtonStyle.red
            super().__init__(label=f"Edit {label}", style=color)

        async def callback(self, interaction: discord.Interaction):
            view: ApprovePunch = self.view
            if interaction.user != view.user:
                await interaction.response.send_message("This is not for you!", ephemeral=True)
                return
            view.value = False
            await interaction.response.send_message("You clicked No!", ephemeral=True)
            message = await view.fetch_message()
            await message.edit(content="You did NOT approve this login attempt.", view=None)
            self.disabled = True

class GetTimeSpent(discord.ui.Modal):
    def __init__(self, button):
        super().__init__(title="Job Completion Form")
        self.button = button
        
        self.text_input = discord.ui.InputText(
            label="Time Spent at Jobsite",
            placeholder="Number of Hours on the qtr hr (e.g. 1, 2.25, 3.5, 4.75)",
            max_length=100,
            style=discord.InputTextStyle.short
        )
        self.add_item(self.text_input)

    async def callback(self, interaction: discord.Interaction):
        user_input = None
        try:
            user_input = float(self.children[0].value)
        except:
            await interaction.respond(f"You entered a non-numerical answer: {user_input}", ephemeral=True)
            user_input = None
        if user_input:
            if (user_input % 0.25 != 0) or (user_input == 0):
                await interaction.respond(f"You entered: {user_input} which is not on the quarter hour or above 0", ephemeral=True)
            else:
                await interaction.respond(f"You entered: `{user_input}` which is a valid response.", ephemeral=True)
                await self.button.handle_modal_response(interaction, user_input)

class CustomerInputModal(discord.ui.Modal):
    def __init__(self, button):
        super().__init__(title="Customer Input")
        self.button = button
        
        self.text_input = discord.ui.InputText(
            label="What customer is this work for?",
            placeholder="Enter customer name",
            max_length=100,
            style=discord.InputTextStyle.short
        )
        self.add_item(self.text_input)

    async def callback(self, interaction: discord.Interaction):
        user_input = self.children[0].value
        await interaction.respond(f"Response recieved successfully to server as: {user_input}", ephemeral=True)
        await self.button.handle_customer_input(interaction, user_input)

class CustomerSelectMenu(discord.ui.Select):
    def __init__(self, options, button):
        self.button = button
        super().__init__(placeholder="Choose a customer", options=options)
        
    async def callback(self, interaction: discord.Interaction):
        customer_id = int(self.values[0])
        await interaction.response.defer()
        await self.button.handle_customer_selection(interaction, customer_id)

async def reloadClockView(user: discord.User, message: discord.Message, bot: discord.Bot, db:str, value:bool = False, currpunch: int = None, ignoreLunchBreak: bool = False):
    new_view = Clock(user=user, message=message, bot=bot, db=db, value=value, currpunch=currpunch, ignoreLunchBreak=ignoreLunchBreak)
    await message.edit(view=new_view)

class Clock(discord.ui.View):
    def __init__(self, user: discord.User, message: discord.Message, bot: discord.Bot, db:str, value:bool = False, currpunch: int = None, ignoreLunchBreak: bool = False):
        super().__init__(timeout=None)
        self.bot: discord.Bot = bot
        self.db: str = db
        self.user: discord.User = user
        self.message: discord.Message = message
        self.value: bool = value
        self.currentpunch: int = currpunch
        self.workpunch: int = None
        self.workpunchType: str = None
        self.ignoreLunchBreak: bool = ignoreLunchBreak

        conn = sqlite3.connect(self.db)
        cursor = conn.cursor()
        if self.currentpunch:
            # Dynamically add buttons based on whether or not we currently have Construction work or service work in progress
            cursor.execute(f"SELECT id, punchType FROM work_time WHERE punchID = {currpunch} AND timeSpent = 0")
            result = cursor.fetchone()
            #print(result)#debugging
            if result:
                self.workpunch = result[0]
                self.workpunchType = result[1]
            
            # Fetch ignoreLunchBreak value
            cursor.execute(f"SELECT ignoreLunchBreak FROM punch_clock WHERE id = {currpunch}")
            lunchBool = cursor.fetchone()[0]
            if lunchBool == "FALSE":
                self.ignoreLunchBreak = False
            elif lunchBool == "TRUE":
                self.ignoreLunchBreak = True
            else:
                self.ignoreLunchBreak = bool(lunchBool)

        # Fetch employee type and allowed work types
        cursor.execute(f"""
            SELECT et.construction, et.service, et.office, e.lunchSkipable 
            FROM employee e 
            JOIN employee_type et ON e.employeeTypeID = et.id 
            WHERE e.id = {self.user.id}
        """)
        emp_type = cursor.fetchone()
        self.lunchSkipable = emp_type[3]
        conn.commit()
        conn.close()
        
        if not self.currentpunch:
            self.add_item(self.ClockInButton())
        elif self.workpunch:
            self.add_item(self.EndWorkPunch(type=self.workpunchType))
            self.add_item(self.EndWorkPunch(custom=True, type=self.workpunchType))
        else:
            self.add_item(self.ClockOutButton())
            if emp_type[0]:  # Construction work allowed
                self.add_item(self.StartWorkPunch(punchType="Construction", stl=discord.ButtonStyle.primary))
            if emp_type[1]:  # Service work allowed
                self.add_item(self.StartWorkPunch(punchType="Service", stl=discord.ButtonStyle.primary))
            if emp_type[2]:  # Office work allowed
                self.add_item(self.StartWorkPunch(punchType="Office", stl=discord.ButtonStyle.primary))
            # Conditionally add Ignore Lunch Break button
            if self.lunchSkipable:
                self.add_item(self.IgnoreLunchBreakButton(self.ignoreLunchBreak))
    
    def get_next_id(self, cursor, table="punch_clock"):
        cursor.execute(f'SELECT MAX(id) FROM {table}')
        result = cursor.fetchone()[0]
        if result is None:
            return 1
        else:
            return result + 1
    
    async def obtain_message(self, bot: discord.bot, channel_id: int, message_id: int):
        cnl = bot.get_channel(channel_id)
        if cnl is None:
            # Fetch the channel if not found in cache
            cnl = await bot.fetch_channel(channel_id)
        return await cnl.fetch_message(message_id)
    
    # clock in
    class ClockInButton(discord.ui.Button):
        def __init__(self):
            super().__init__(label=f"Clock-In", style=discord.ButtonStyle.green)
        
        async def callback(self, interaction: discord.Interaction):
            user = interaction.user
            role = None
            passedChecks = True
            if os.getenv('TIMECARD_TIMECLOCK_ROLE_ID'):
                role = discord.utils.get(user.roles, id=int(os.getenv('TIMECARD_TIMECLOCK_ROLE_ID')))
            if interaction.user != self.view.user and not canPressButton(interaction, self.view.user, ['TIMECARD_ADMIN_ROLE', 'TIMECARD_TIMECLOCK_ROLE_ID']):
                await interaction.response.send_message("This is not for you!", ephemeral=True)
                passedChecks = False
            if self.view.value == True:
                await interaction.response.send_message("You are already clocked in!", ephemeral=True)
                passedChecks = False
            if passedChecks:
                #connect to the db
                conn = sqlite3.connect(self.view.db)
                cursor = conn.cursor()
                #set the values for clocking in
                now = datetime.now()
                now_str = now.strftime('%Y-%m-%d %H:%M:%S')
                #check if the person who has pressed the button is allowed to clock in w/o approval
                punchInApproval = True
                approvalMessage: discord.Message = None
                nextid = self.view.get_next_id(cursor)
                values = (nextid, user.id, now_str, punchInApproval, None, None)
                if role is None:
                    punchInApproval = False
                    approvalMessage = await self.view.bot.get_channel(int(os.getenv('TIMECARD_ADMIN_CHANNEL_ID'))).send(f"<@{self.view.user.id}> attempted to login today at {now_str} in a non-standard way.\nDo you approve of this login attempt?")
                    #values = (id, employeeID, punchInTime, punchInApproval, checkChannelId, checkMessageId) for reference
                    values = (nextid,user.id, now_str, punchInApproval, approvalMessage.channel.id, approvalMessage.id)
                ## store the clock-in in the db
                cursor.execute(f'INSERT INTO punch_clock (id, employeeID, punchInTime, punchInApproval, checkChannelId, checkMessageId) VALUES (?, ?, ?, ?, ?, ?)', values)
                ## wrap it all up with a nice clean bow
                conn.commit()
                conn.close()
                if not punchInApproval:
                    newview = ApprovePunch(punch=nextid, message=approvalMessage, bot=self.view.bot, db=self.view.db)
                    await approvalMessage.edit(view=newview)
                #update the embed
                self.view.value = True
                embeds = self.view.message.embeds
                embeds[0].color = discord.Colour.brand_green()
                embeds[0].title = "You ARE currently clocked in."
                embeds[0].set_footer(text = str(nextid))
                self.view.currentpunch = nextid
                await self.view.message.edit(embeds=embeds)
                userStr = ""
                if interaction.user.id != self.view.user.id:
                    userStr = f" by {interaction.user}"
                print(f"{self.view.user} just clocked in{userStr}.")
                if interaction.user.id != self.view.user.id:
                    userStr = f" {self.view.user}"
                await interaction.response.send_message(f"You clocked in{userStr}.", ephemeral=True)
                await reloadClockView(user=self.view.user, message=self.view.message, bot=self.view.bot, db=self.view.db, value=self.view.value, currpunch=self.view.currentpunch, ignoreLunchBreak=self.view.ignoreLunchBreak)

    # start work punch
    class StartWorkPunch(discord.ui.Button):
        def __init__(self, punchType: str, stl: discord.ButtonStyle):
            self.punchType = punchType
            super().__init__(label=f"Start {punchType}", style=stl)
        
        async def callback(self, interaction: discord.Interaction):
            passedChecks = True
            if interaction.user != self.view.user and not canPressButton(interaction, self.view.user, ['TIMECARD_ADMIN_ROLE', 'TIMECARD_TIMECLOCK_ROLE_ID']): 
                await interaction.response.send_message("This is not for you!", ephemeral=True)
                passedChecks = False
            if self.view.value == False:
                await interaction.response.send_message("You are currently clocked out and can't start work until you clock in!", ephemeral=True)
                passedChecks = False
            if passedChecks:
                if self.punchType == "Office":
                    await interaction.response.defer()
                    await self.handle_customer_selection(interaction=interaction, customer_id=0)
                else:
                    modal = CustomerInputModal(button=self)
                    await interaction.response.send_modal(modal)
        
        async def handle_customer_input(self, interaction: discord.Interaction, user_input: str):
            conn = sqlite3.connect(self.view.db)
            cursor = conn.cursor()
            cursor.execute(f"SELECT id, name FROM customer WHERE name LIKE ? ORDER BY name LIMIT 25", (f'%{user_input}%',))
            results = cursor.fetchall()
            conn.close()

            if results:
                options = [discord.SelectOption(label=name, value=str(customer_id)) for customer_id, name in results]
                select_menu = CustomerSelectMenu(options=options, button=self)
                view = discord.ui.View()
                view.add_item(select_menu)
                await interaction.followup.send("Select a customer:", view=view, ephemeral=True)
            else:
                await interaction.followup.send("No customers found.", ephemeral=True)

        async def handle_customer_selection(self, interaction: discord.Interaction, customer_id: int):
            conn = sqlite3.connect(self.view.db)
            cursor = conn.cursor()
            next_id = self.view.get_next_id(cursor, "work_time")
            time_started = datetime.now().isoformat()
            cursor.execute(f"SELECT name FROM customer WHERE id = {customer_id}")
            customerName = cursor.fetchone()[0]
            cursor.execute("INSERT INTO work_time (id, punchID, customerID, punchType, timeStarted) VALUES (?, ?, ?, ?, ?)", 
                        (next_id,self.view.currentpunch, customer_id, self.punchType, time_started))
            conn.commit()
            conn.close()
            await interaction.followup.send(f"Work punch started for customer ID {customer_id} ({customerName}).", ephemeral=True)
            self.view.workpunch = next_id
            self.view.workpunchType = self.punchType
            await reloadClockView(user=self.view.user, message=self.view.message, bot=self.view.bot, db=self.view.db, value=self.view.value, currpunch=self.view.currentpunch, ignoreLunchBreak=self.view.ignoreLunchBreak)
    
    # end work punch
    class EndWorkPunch(discord.ui.Button):
        def __init__(self, custom:bool = False, type: str = "Construction"):
            self.timeSpent: float = None
            self.custom = custom
            txt = "Now"
            stl = discord.ButtonStyle.primary
            if custom:
                txt = "Custom"
                stl = discord.ButtonStyle.secondary
            super().__init__(label=f"End {type} Work {txt}", style=stl)
        
        async def callback(self, interaction: discord.Interaction):
            if interaction.user != self.view.user and not canPressButton(interaction, self.view.user, ['TIMECARD_ADMIN_ROLE', 'TIMECARD_TIMECLOCK_ROLE_ID']): 
                await interaction.response.send_message("This is not for you!", ephemeral=True)
            else:
                if self.custom:
                    modal = GetTimeSpent(button=self)
                    await interaction.response.send_modal(modal)
                else:
                    conn = sqlite3.connect(self.view.db)
                    cursor = conn.cursor()
                    cursor.execute(f"SELECT timeStarted FROM work_time WHERE id = {self.view.workpunch}")
                    result = cursor.fetchone()
                    conn.commit()
                    conn.close()
                    if result:
                        time_started = datetime.fromisoformat(result[0])
                        time_now = datetime.now()
                        time_difference = time_now - time_started
                        
                        # Convert the time difference to hours
                        hours_spent = time_difference.total_seconds() / 3600
                        
                        # Round to the nearest quarter hour
                        nearest_quarter_hour = round(hours_spent * 4) / 4
                        if nearest_quarter_hour == 0:
                            nearest_quarter_hour = 0.25
                        print(f"Time spent: {nearest_quarter_hour} hours.")
                        self.timeSpent = nearest_quarter_hour
                        await self.completeCallbackMethod(interaction=interaction)
                    else:
                        await interaction.response.send_message("Start time not found.", ephemeral=True)
            
        async def completeCallbackMethod(self, interaction: discord.Interaction):
            conn = sqlite3.connect(self.view.db)
            cursor = conn.cursor()
            print(f"UPDATE work_time SET timeSpent = {int(self.timeSpent * 60)} WHERE id = {self.view.workpunch}")#debugging
            cursor.execute(f"UPDATE work_time SET timeSpent = {int(self.timeSpent * 60)} WHERE id = {self.view.workpunch}")
            conn.commit()
            conn.close()
            self.view.workpunch = None
            self.view.workpunchType = None
            await interaction.respond(f"You have successfully completed your work at the jobsite in {self.timeSpent} hour(s).", ephemeral=True)
            # Switch out the buttons
            await reloadClockView(user=self.view.user, message=self.view.message, bot=self.view.bot, db=self.view.db, value=self.view.value, currpunch=self.view.currentpunch, ignoreLunchBreak=self.view.ignoreLunchBreak)
        
        async def handle_modal_response(self, interaction: discord.Interaction, time_spent: float):
            self.timeSpent = time_spent
            await self.completeCallbackMethod(interaction=interaction)

    # clock out
    class ClockOutButton(discord.ui.Button):
        def __init__(self):
            super().__init__(label=f"Clock-Out", style=discord.ButtonStyle.red)
        
        async def callback(self, interaction: discord.Interaction):
            user = interaction.user
            role = None
            passedChecks = True
            if os.getenv('TIMECARD_TIMECLOCK_ROLE_ID'):
                role = discord.utils.get(user.roles, id=int(os.getenv('TIMECARD_TIMECLOCK_ROLE_ID')))
            if interaction.user != self.view.user and not canPressButton(interaction, self.view.user, ['TIMECARD_ADMIN_ROLE', 'TIMECARD_TIMECLOCK_ROLE_ID']): 
                await interaction.response.send_message("This is not for you!", ephemeral=True)
                passedChecks = False
            if self.view.value == False:
                await interaction.response.send_message("You are already clocked out!", ephemeral=True)
                passedChecks = False
            if passedChecks:
                #variables used throughout
                punchOutApproval = True
                approvalMessage: discord.Message = None
                content: str = None
                #connect to the db
                conn = sqlite3.connect(self.view.db)
                cursor = conn.cursor()
                currPunch = None
                if self.view.currentpunch:
                    currPunch = self.view.currentpunch
                else:
                    cursor.execute(f'SELECT MAX(id) FROM punch_clock WHERE employeeID = {self.view.user.id} AND punchOutTime is NULL ')
                    currPunch = cursor.fetchone()[0]
                if currPunch is None:
                    print("Critical error: I dont think we shouldve been able to clock out if the employee was never clocked in!")
                cursor.execute(f'SELECT checkChannelId, checkMessageId FROM punch_clock WHERE id = {currPunch}')
                result = cursor.fetchone()
                if not(result[0] is None and result[1] is None):
                    approvalMessage = await self.view.obtain_message(bot=self.view.bot, channel_id=int(result[0]), message_id=int(result[1]))
                #set the values for clocking in
                now = datetime.now()
                now_str = now.strftime('%Y-%m-%d %H:%M:%S')
                #check if the person who has pressed the button is allowed to clock in w/o approval
                if role is None:
                    punchOutApproval = False
                    cursor.execute(f'SELECT checkChannelId, checkMessageId FROM punch_clock WHERE id = {currPunch}')
                    result = cursor.fetchone()
                    if result[0] is None and result[1] is None:
                        approvalMessage = await self.view.bot.get_channel(int(os.getenv('TIMECARD_ADMIN_CHANNEL_ID'))).send(f"<@{self.view.user.id}> attempted to logout today at {now_str} in a non-standard way.\nDo you approve of this login attempt?")
                    else:
                        content = approvalMessage.content
                        content = content[:-37:] + f"<@{self.view.user.id}> attempted to logout today at {now_str} in a non-standard way.\n" + content[-37::]
                ## store the clock-out in the db
                if approvalMessage:
                    print(f'EXECUTING COMMAND:\nUPDATE punch_clock SET punchOutTime = "{now_str}", punchOutApproval = {punchOutApproval}, checkChannelId = {approvalMessage.channel.id}, checkMessageId = {approvalMessage.id} WHERE id = {currPunch}')#debugging
                    cursor.execute(f'UPDATE punch_clock SET punchOutTime = "{now_str}", punchOutApproval = {punchOutApproval}, checkChannelId = {approvalMessage.channel.id}, checkMessageId = {approvalMessage.id} WHERE id = {currPunch}')
                else:
                    print(f'EXECUTING COMMAND:\nUPDATE punch_clock SET punchOutTime = "{now_str}", punchOutApproval = {punchOutApproval}, checkChannelId = NULL, checkMessageId = NULL WHERE id = {currPunch}')#debugging
                    cursor.execute(f'UPDATE punch_clock SET punchOutTime = "{now_str}", punchOutApproval = {punchOutApproval}, checkChannelId = NULL, checkMessageId = NULL WHERE id = {currPunch}')
                ## wrap it all up with a nice clean bow
                conn.commit()
                conn.close()
                if approvalMessage:
                    view = ApprovePunch(punch=currPunch, message=approvalMessage, bot=self.view.bot, db=self.view.db)
                    if content:
                        await approvalMessage.edit(content=content, view=view)
                    else:
                        await approvalMessage.edit(view=view)
                #update the embed
                self.view.value = False
                embeds = self.view.message.embeds
                embeds[0].color = discord.Colour.brand_red()
                embeds[0].title = "You are currently NOT clocked in."
                embeds[0].set_footer(text = f"User: {self.view.user.name}")
                self.view.currentpunch = None
                await self.view.message.edit(embeds=embeds)
                userStr = ""
                if interaction.user.id != self.view.user.id:
                    userStr = f" by {interaction.user}"
                print(f"{self.view.user} just clocked out{userStr}.")
                if interaction.user.id != self.view.user.id:
                    userStr = f" {self.view.user}"
                await interaction.response.send_message(f"You clocked out{userStr}.", ephemeral=True)
                await reloadClockView(user=self.view.user, message=self.view.message, bot=self.view.bot, db=self.view.db, value=self.view.value, currpunch=self.view.currentpunch, ignoreLunchBreak=self.view.ignoreLunchBreak)
    
    # toggle whether to ignore lunch break or not
    class IgnoreLunchBreakButton(discord.ui.Button):
        def __init__(self, ignoreLunchBreak):
            label = "Ignoring Lunch Break" if ignoreLunchBreak else "NOT Ignoring Lunch Break"
            style = discord.ButtonStyle.success if ignoreLunchBreak else discord.ButtonStyle.danger
            super().__init__(label=label, style=style)
        
        async def callback(self, interaction: discord.Interaction):
            if not canPressButton(interaction, self.view.user, ['TIMECARD_ADMIN_ROLE', 'TIMECARD_TIMECLOCK_ROLE_ID']):
                await interaction.response.send_message("You don't have permission to use this button.", ephemeral=True)
                return
            conn = sqlite3.connect(self.view.db)
            cursor = conn.cursor()
            new_value = not self.view.ignoreLunchBreak
            cursor.execute(f"UPDATE punch_clock SET ignoreLunchBreak = ? WHERE id = ?", (new_value, self.view.currentpunch))
            conn.commit()
            conn.close()
            
            self.view.ignoreLunchBreak = new_value
            self.style = discord.ButtonStyle.success if new_value else discord.ButtonStyle.danger
            self.label = "Ignoring Lunch Break" if new_value else "NOT Ignoring Lunch Break"
            await interaction.response.edit_message(view=self.view)


class TimeTracking(commands.Cog): # create a class for our cog that inherits from commands.Cog
    # this class is used to create a cog, which is a module that can be added to the bot

    def __init__(self, bot): # this is a special method that is called when the cog is loaded
        self.cwd = os.getcwd()
        self.db = self.cwd + "/timetracker.db"
        self.dbSetup(self.db)
        self.bot = bot

    # Gets all of the view buttons working again after a bot restart
    @commands.Cog.listener()
    async def on_ready(self):
        # Connect to your SQLite database
        conn = sqlite3.connect(self.db)
        cursor = conn.cursor()
        print("Re-initializing Clock Views for each Employee...")
        cursor.execute("SELECT id, clockChannelId, clockMessageId FROM employee WHERE clockChannelId is not NULL AND clockMessageId is not NULL")
        for user_data in cursor.fetchall():
            print(f"Employee: {user_data}")
            userObj = await self.bot.fetch_user(user_data[0])
            msg: discord.Message = None
            chnl = self.bot.get_channel(user_data[1])
            if chnl is not None:
                msg = await chnl.fetch_message(user_data[2])
            else:
                print(f"Channel with ID {user_data[2]} not found.")
            if msg is None:
                print(f"Failed to get Clock for user: {user_data}")
            else:
                clockedIn = False
                currentPunch = None
                if msg.embeds[0].title == "You ARE currently clocked in.":
                    clockedIn = True
                    currentPunch = int(msg.embeds[0].footer.text)
                new_view = Clock(user=userObj, message=msg, bot=self.bot, db=self.db, value=clockedIn, currpunch=currentPunch)
                await msg.edit(view=new_view)
        print("Finished re-initializing Clock Views for each Employee...")
        print("Re-initializing Approval Message Views for each Punch...")
        cursor.execute("SELECT id, checkChannelId, checkMessageId FROM punch_clock WHERE checkChannelId is not NULL AND checkMessageId is not NULL")
        for punch_data in cursor.fetchall():
            print(f"Punch: {punch_data}")
            msg: discord.Message = None
            chnl = self.bot.get_channel(punch_data[1])
            if chnl is not None:
                msg = await chnl.fetch_message(punch_data[2])
            else:
                print(f"Channel with ID {punch_data[2]} not found.")
            if msg is None:
                print(f"Failed to get Punch Approval Message for punch: {punch_data}")
            else:
                new_view = ApprovePunch(punch=punch_data[0], message=msg, bot=self.bot, db=self.db)
                await msg.edit(view=new_view)
        print("Finished re-initializing Approval Message Views for each Punch...")

                    ## Employee Type Methods
        ### Less important methods ###
    #add type(name, rate, id=-1) -  admin command
    #edit type(id, name="", rate="") -  admin command
    #remove type(id) - admin command

    
    @discord.slash_command(name="addcustomer", description="Add a new Customer to the customer table.")
    @commands.has_permissions(administrator=True)
    async def addcustomer(self,
                          ctx: discord.ApplicationContext,
                          name: discord.Option(str, description="Full Name of Customer")   # type: ignore
                          ):
        conn = sqlite3.connect(self.db)
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM customer WHERE name = '{name}'")
        customers = cursor.fetchall()
        if not customers:
            print(f"EXECUTING COMMAND:\nINSERT INTO customer (name) VALUES ({name})")
            cursor.executemany(f'INSERT INTO customer (name) VALUES (?)',[(name,)])
            await ctx.respond(f"Successfully inserted {name} into customer table", ephemeral=True)
        else:
            await ctx.respond(f"Could not insert {name} into customer table because it already exists at:\n{customers}", ephemeral=True)
        conn.commit()
        conn.close()

    @discord.slash_command(name="editcustomer", description="Edit an existing Customer in the customer table.")
    @commands.has_permissions(administrator=True)
    async def editcustomer(self,
                          ctx: discord.ApplicationContext,
                          newname: discord.Option(str, description="New name for Customer"),   # type: ignore
                          id: discord.Option(int, default=None, description="Id of Customer"),   # type: ignore
                          name: discord.Option(str, default=None, description="Name of Customer")   # type: ignore
                          ):
        conn = sqlite3.connect(self.db)
        cursor = conn.cursor()
        if id is None and name is None:
            await ctx.respond(f"Can't have both id and name parameter be empty must use one... \nIf you put both in, it will prioritize id over name.", ephemeral=True)
        elif id:
            try:
                cursor.execute(f"SELECT * FROM customer WHERE id = {id}")
                customers = cursor.fetchall()
                if not customers:
                    await ctx.respond(f"Could not find customer: {name}", ephemeral=True)
                else:
                    print(f"EXECUTING COMMAND:\nUPDATE customer SET name = {newname} WHERE id = {customers[0][0]}")
                    cursor.executemany(f'UPDATE customer SET name = ? WHERE id = ?',[(newname, customers[0][0]),])
                    await ctx.respond(f"Successfully updated {customers[0][0]}, {customers[0][1]} in the customer table to {newname}", ephemeral=True)
            except:
                await ctx.respond(f"Invalid ID paramter", ephemeral=True)
        else:
            try:
                cursor.execute(f"SELECT * FROM customer WHERE name = '{name}'")
                customers = cursor.fetchall()
                if not customers or (len(customers) > 1):
                    await ctx.respond(f"Could not get customer to: {name} because the search returned:\n{customers}", ephemeral=True)
                else:
                    print(f"EXECUTING COMMAND:\nUPDATE customer SET name = {newname} WHERE id = {customers[0][0]}")
                    cursor.executemany(f'UPDATE customer SET name = ? WHERE id = ?',[(newname, customers[0][0]),])
                    await ctx.respond(f"Successfully updated {customers[0][0]}, {customers[0][1]} in the customer table to {newname}", ephemeral=True)
            except:
                await ctx.respond(f"Invalid name paramter", ephemeral=True)
        conn.commit()
        conn.close()

    #
                    ## Employee Methods
    async def employee_type_autocomplete(ctx: discord.AutocompleteContext):
        conn = sqlite3.connect(os.getcwd() + "/timetracker.db")
        cursor = conn.cursor()
        cursor.execute("SELECT id, name FROM employee_type")
        types = cursor.fetchall()
        conn.close()
        return [discord.OptionChoice(name=type[1], value=type[0]) for type in types]
    

    # add employee(name, phonenumber, addressline1, city, state, zip, addressline2="",user="") - admin command   *updating needed
    @discord.slash_command(name="addemployee", description="Add a new Employee to from discord to the system.")
    @commands.has_permissions(administrator=True)
    async def addemployee(self, 
                   ctx: discord.ApplicationContext,
                   name: discord.Option(str, description="Full Name of Employee"),   # type: ignore
                   phonenumber: discord.Option(str, description="Phone Number of Employee"),   # type: ignore
                   addressline1: discord.Option(str, description="Address Line 1 of Employee"),   # type: ignore
                   city: discord.Option(str, description="Address City of Employee"),   # type: ignore
                   state: discord.Option(str, description="Address State of Employee"),   # type: ignore
                   zip: discord.Option(str, description="Address Zip of Employee"),   # type: ignore
                   addressline2: discord.Option(str, default="", description="Address Line 2 of Employee"),   # type: ignore
                   user: discord.Option(str, default=None, description="Add a different user to the employee table"),   # type: ignore
                   payrate: discord.Option(float, default=16.00, description="Payrate for the Employee"),   # type: ignore
                   employeetype: discord.Option(int, default=2, description="Employee Type", autocomplete=employee_type_autocomplete),   # type: ignore
                   ):
        conn = sqlite3.connect(self.db)
        cursor = conn.cursor()
        id = ctx.author.id
        valid = True
        if user is not None:
            try:
                newid = int(user.strip()[2:-1:])
                id = newid
            except:
                valid = False
                await ctx.respond("User Override Attribute was improperly formatted.")
                print(ctx.author + "User Override Attribute was improperly formatted.")
        
        cursor.execute(f"SELECT id FROM employee WHERE id = {id}")
        employees = cursor.fetchall()
        if len(employees) > 0:
            valid = False
            await ctx.respond(f"Cannot add new employee {name} (aka <@{id}>) because they already exist in the database.")
            print(f"Cannot add new employee {name} (aka <@{id}>) because they already exist in the database.")
        if valid:
            value = [(id, name, phonenumber, addressline1, addressline2, city, state, zip, payrate, employeetype)]
            try:
                cursor.executemany('INSERT INTO employee (id, name, phoneNumber, addressLine1, addressLine2, addressCity, addressState, addressZip, payrate, employeeTypeID) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)', value)
                conn.commit()
                conn.close()
                await ctx.respond(f"Added new employee {value[0][1]} (aka <@{value[0][0]}>)")
                print(f"Added new employee {value[0][1]} (aka <@{value[0][0]}>)")
            except:
                await ctx.respond(f"Error adding new employee {value[0][1]} (aka <@{value[0][0]}>)")
                print(f"Error adding new employee {value[0][1]} (aka <@{value[0][0]}>) by {ctx.author}")
    # Error handling for missing permissions
    @addemployee.error
    async def addemployee_error(self, ctx, error):
        if isinstance(error, commands.MissingPermissions):
            await ctx.respond(f"{ctx.author}, you do not have the necessary permissions to use this command.", ephemeral=True)
            print(f"{ctx.author}, you do not have the necessary permissions to use this command.")
        
    #edit employee(user, name="", phonenumber="", addressline1="", addressline2="", city="", state="",, zip="") - admin command
    #remove employee(user) - admin command

                ## Punch Clock Methods
    # create clock(user, channelid="") - admin command
    @discord.slash_command(name="createclock", description="Create a time clock embed for a user.")
    @commands.has_permissions(administrator=True)
    async def createclock(self, 
                          ctx: discord.ApplicationContext,
                          user: discord.Option(str, description="The discord user you wish to make a time clock for."),   # type: ignore
                          channel: discord.Option(str, default=None, description="Choose a different channel to create the time clock.")   # type: ignore
                          ):
        ## check that the values given were in the correct format
        try:
            _ = int(user[2:-1:])
            if channel is not None:
                _ = int(user[2:-1:])
        except:
            if channel is None:
                await ctx.respond(f"The parameters submitted were not in the correct format for {user}: {user[2:-1:]}", ephemeral=True)
            else:
                await ctx.respond(f"The parameters submitted were not in the correct format for {user}: {user[2:-1:]} or {channel[2:-1:]}", ephemeral=True)
            return
        # connect to db
        conn = sqlite3.connect(self.db)
        cursor = conn.cursor()
        ## use the data in the db to check that we aren't overwriting good data
        cursor.execute(f"SELECT clockChannelId, clockMessageId FROM employee WHERE id = {user[2:-1:]}")
        employees = cursor.fetchall()
        if len(employees) == 0:
            await ctx.respond(f"Cannot edit employee {user} because they don't exist in the database.", ephemeral=True)
            print(f"Cannot edit employee {user} because they don't exist in the database.")
            conn.commit()
            conn.close()
            return
        elif len(employees) > 1:
            await ctx.respond(f"There is a significant bug in the code that <@336231886198276096> needs to look into.", ephemeral=True)
            print("How did we get multiple employees back with the same id?!!!!!")
            conn.commit()
            conn.close()
            return
        elif employees[0][1] is not None:
            view = Confirm(user=ctx.user, timeout=180)
            await ctx.response.send_message("This emplyee already has a time clock, do you want to proceed?", view=view, ephemeral=True)
            await view.wait()
            if view.value:
                oldMessageID = employees[0][1]
                oldChannelID = employees[0][0]
                print(f"Override Confirmed by {view.user}... Trying to delete old time clock message.")
                await self.deleteclock(ctx, user, oldChannelID, oldMessageID)
            else:
                await ctx.followup.send("You chose not to proceed or did not respond in time.", ephemeral=True)
                return
        ##figure out if the user is clocked in already
        clockedIn = False
        cursor.execute(f"SELECT id, employeeID, punchOutTime FROM punch_clock WHERE employeeID = {user[2:-1:]} ORDER BY id DESC LIMIT 1")
        punches = cursor.fetchall()
        currentPunch = None
        if len(punches) > 0:
            if not punches[0][2]:
                clockedIn = True
                currentPunch = punches[0][0]
        ## go to channel and send the message
        channelObj = ctx.channel
        if channel is not None:
            channelObj = self.bot.get_channel(int(channel[2:-1:]))
        userObj = await self.bot.fetch_user(int(user[2:-1:]))
        if not userObj:
            await ctx.respond("Failed to find member?!", ephemeral=True)
            return
        ##create the embed
        embed = discord.Embed(
            title="You are currently NOT clocked in.",
            color=discord.Colour.brand_red(), # Pycord provides a class with default colors you can choose from
        )
        embed.add_field(name="Wondering how to clock in?", value="Click the green clock-in button and watch the field turn green. And to clock out, hit the red clock-out button. Simple as that!")
        embed.set_footer(text=f"User: {userObj.name}")
        embed.set_author(name=f"{userObj} Time Clock", icon_url="https://media.discordapp.net/attachments/1224574847213109330/1244848933675728978/clkfbambooblack600x600-bgf8f8f8.png?ex=66569b69&is=665549e9&hm=61d63652381160c0339fe9fb51cceb8d6971c1878e01b9cb0a181d43e5d97546&=&format=webp&quality=lossless")
        message = await channelObj.send(embed=embed)
        #if already clocked in (in the database) then we gotta make sure this newly created punch clock starts on the right value
        if clockedIn:
            embeds = message.embeds
            embeds[0].color = discord.Colour.brand_green()
            embeds[0].title = "You ARE currently clocked in."
            embeds[0].set_footer(text = str(currentPunch))
            await message.edit(embeds=embeds)
        view = Clock(user=userObj, message=message, bot=self.bot, db=self.db, value=clockedIn, currpunch=currentPunch)
        await message.edit(view=view)
        ## store the message id and channel id in the db
        cursor.execute(f"UPDATE employee SET clockChannelId = {message.channel.id}, clockMessageId = {message.id} WHERE id = {user[2:-1:]}")
        ## wrap it all up with a nice clean bow
        conn.commit()
        conn.close()
        await ctx.respond(f"Clock created successfully for {user}", ephemeral=True)
        print(f"{ctx.author} - Clock created successfully for {user}")

        
    # delete clock(user) - admin command
    @discord.slash_command(name="deleteclock", description="Delete a clock embed for a user.")
    @commands.has_permissions(administrator=True)
    async def deleteclock(self,
                          ctx: discord.ApplicationContext,
                          user: discord.Option(str, description="The discord user you wish to delete their time clock for"), # type: ignore
                          channelid: discord.Option(int, default=None, description="The discord channelId where the msg you wish to delete is"), # type: ignore
                          messageid: discord.Option(int, default=None, description="The discord msgId where the msg you wish to delete is"), # type: ignore
                          ):
        ctxInitiated = True
        successful = True
        employees = None
        oldChannelID = channelid
        oldMessageID = messageid
        if oldChannelID and oldMessageID:
            ctxInitiated = False
        if ctxInitiated:
            # connect to db
            conn = sqlite3.connect(self.db)
            cursor = conn.cursor()
            ## use the data in the db to check that we aren't overwriting good data
            cursor.execute(f"SELECT clockChannelId, clockMessageId FROM employee WHERE id = {user[2:-1:]}")
            employees = cursor.fetchall()
            oldMessageID = employees[0][1]
            oldChannelID = employees[0][0]
        try:
            # Fetch the channel by ID
            delchannel = self.bot.get_channel(oldChannelID)
            if delchannel is not None:
                message = await delchannel.fetch_message(oldMessageID)
                # Delete the message
                await message.delete()
                print(f"Message with ID {oldMessageID} deleted from channel {oldChannelID}.")
            else:
                print(f"Channel with ID {oldChannelID} not found.")
        except discord.NotFound:
            print("Message or channel not found.")
            successful = False
        except discord.Forbidden:
            print("I do not have permission to delete this message.")
            successful = False
        except discord.HTTPException as e:
            print(f"Failed to delete message: {e}")
            successful = False
        if ctxInitiated and successful:
            await ctx.respond(f"Clock deleted successfully for user: {user}!")


        ### Less important methods ###
    
    # display punches(user="")
    #edit punch(id) - admin command - trigger via button?


                ## Work Time Methods
    # display work time(user)

    #add work time(customer, punchtype, timespent, punchid=-1)

        ### Less important methods ###
    #edit work time(id, punchid=-1, customer="", punchtype="", timespent=0) - admin perms to edit someone else's - trigger via button?
    #remove work time(id) - admin command? - trigger via button?

                ## Reports
    async def employee_group_autocomplete(self, ctx: discord.AutocompleteContext):
        conn = sqlite3.connect(self.db)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM employee_group")
        groups = cursor.fetchall()
        conn.close()
        return [group[0] for group in groups if ctx.value.lower() in group[0].lower()]

    def round_to_quarter_hour(self, minutes):
        return round(minutes / 15) * 15

    def convert_minutes_to_hours(self, minutes):
        return minutes / 60

    def is_saturday(self, date_str):
        # Parse the date string into a datetime object
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        # Check if the day of the week is Saturday (5)
        return date_obj.weekday() == 5
    
    def get_day_of_week(self, date_str):
        # Parse the date string into a datetime object
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        
        # Get the day of the week as a full name (e.g., "Monday", "Tuesday")
        day_of_week = date_obj.strftime('%A')
        
        return day_of_week

    def setCell(self, cell, newValue, newFont, newNumberFormat, newBorder, newAlignment):
        cell.value = newValue
        cell.font = newFont
        cell.number_format = newNumberFormat
        cell.border = newBorder
        cell.alignment = newAlignment

    def getKey(self, e):
        dat1, dat2 = e
        return datetime.strptime(dat1[2], "%Y-%m-%d %H:%M:%S")

    def createReportWorkbook(self, new_wb, template_sheet_name):
        # Load the existing workbook
        template_path = 'templates/Template Sheets.xlsx'
        newWorkbook = xwriter.Workbook(new_wb)
        newWorkbook.add_worksheet(template_sheet_name)
        newWorkbook.close()
        # Copy the template sheet to the new workbook
        tpwb = load_workbook(template_path)
        newWbObj = load_workbook(new_wb)
        tpws = tpwb[template_sheet_name]
        newWsObj =newWbObj[template_sheet_name]
        for columns in tpws.iter_cols(min_row=1, min_col=1, max_col=tpws.max_column, max_row=1):
            for col in columns:
                col = col.coordinate[:-1:]
                newWsObj.column_dimensions[col].width = tpws.column_dimensions[col].width
        for columns in tpws.iter_cols(min_row=1, min_col=1, max_col=tpws.max_column, max_row=tpws.max_row):
            for cell in columns:
                self.setCell(newWsObj[f"{cell.coordinate}"], cell.value, copy(cell.font), copy(cell.number_format), copy(cell.border), copy(cell.alignment))
        newWbObj.save(new_wb)

            

    # This method takes the punch data for one employee and formats the given openpyxl sheet with the data
    def reportTimecardData(
        self,
        sheet: openpyxl.worksheet,
        data: list
    ):
        #last row used so that we can set the print_area correctly and will be returned at the end of the method
        lastRow = 4
        #sort the data in order by punch_in datetime (chronilogical order) then store it in a dictionary
        data.sort(key=self.getKey)
        databyDate = {}
        for tup in data:
            dataDate = datetime.strptime(tup[0][2], '%Y-%m-%d %H:%M:%S').date()
            if dataDate not in databyDate.keys():
                databyDate[dataDate] = [tup]
            else:
                dateLst = databyDate[dataDate]
                dateLst.append(tup)
                databyDate[dataDate] = dateLst
        #totals
        TotalConstructionTime = TotalServiceTime = 0
        anyUnapprovedPunches = False
        ##all of my default styling with fonts, number formats, borders, and alignment
        normal_font, bold_font = Font(name='Arial', size=10, bold=False), Font(name='Arial', size=10, bold=True)
        date_number_format, hrs_number_format, time_duration_format, time_format = 'mm-dd-yy', '0.00 "hrs"', 'h:mm:ss', 'hh:mm AM/PM'
        workTime_border = Border( top=Side(border_style='thin', color='00000000'), bottom=Side(border_style='thin', color='00000000') )
        rightAlign, centerAlign, leftAlign = Alignment(horizontal='right'), Alignment(horizontal='center'), Alignment(horizontal='left')
        #time to get data onto the page
        for date in databyDate.keys():
            TotalingRow = lastRow
            TotalTime = TotOfficeTime = TotShopTime = TotLunchTime = 0
            allPunchesOnDate = databyDate[date]
            for punch, work in allPunchesOnDate:
                punch_in = datetime.strptime(punch[2], '%Y-%m-%d %H:%M:%S')
                punch_out = datetime.strptime(punch[3], '%Y-%m-%d %H:%M:%S') if punch[3] else datetime.now()
                shift_duration = self.convert_minutes_to_hours(self.round_to_quarter_hour((punch_out - punch_in).total_seconds() / 60))
                lunch = 0.5 if not punch[6] and shift_duration >= 6.0 else 0
                constructionLst = []
                serviceLst = []
                TotConst = TotSer = TotOfc = 0
                if work:
                    for worktype, cust, time in work:
                        hrs = time / 60
                        if worktype == "Office":
                            TotOfc = TotOfc + hrs
                            TotOfficeTime = TotOfficeTime + hrs
                        elif worktype == "Service":
                            TotSer = TotSer + hrs
                            serviceLst.append((cust, hrs))
                        else:
                            TotConst = TotConst + hrs
                            constructionLst.append((cust, hrs))
                #start inserting punch data into sheet
                lastRow = lastRow + 1
                self.setCell(sheet[f"I{lastRow}"], "Clock-In", bold_font, 'General', Border(), Alignment())
                self.setCell(sheet[f"J{lastRow}"], punch_in.time(), normal_font, time_format, Border(), leftAlign)
                if punch[4]:
                    self.setCell(sheet[f"K{lastRow}"], "<- Unapproved Punch", normal_font, 'General', Border(), leftAlign)
                    anyUnapprovedPunches = True
                if len(constructionLst) > 0:
                    lastRow = lastRow + 1
                    self.setCell(sheet[f"J{lastRow}"], "Construction", bold_font, 'General', workTime_border, rightAlign)
                    self.setCell(sheet[f"K{lastRow}"], TotConst, normal_font, hrs_number_format, workTime_border, Alignment())
                    TotalConstructionTime = TotalConstructionTime + TotConst
                    for name, hrs in constructionLst:
                        lastRow = lastRow + 1
                        self.setCell(sheet[f"J{lastRow}"], name, normal_font, 'General', Border(), rightAlign)
                        self.setCell(sheet[f"K{lastRow}"], hrs, normal_font, hrs_number_format, Border(), Alignment())
                if len(serviceLst) > 0:
                    lastRow = lastRow + 1
                    self.setCell(sheet[f"J{lastRow}"], "Service", bold_font, 'General', workTime_border, rightAlign)
                    self.setCell(sheet[f"K{lastRow}"], TotSer, normal_font, hrs_number_format, workTime_border, Alignment())
                    TotalServiceTime = TotalServiceTime + TotSer
                    for name, hrs in serviceLst:
                        lastRow = lastRow + 1
                        self.setCell(sheet[f"J{lastRow}"], name, normal_font, 'General', Border(), rightAlign)
                        self.setCell(sheet[f"K{lastRow}"], hrs, normal_font, hrs_number_format, Border(), Alignment())
                lastRow = lastRow + 1
                self.setCell(sheet[f"I{lastRow}"], "Clock-Out", bold_font, 'General', Border(), Alignment())
                self.setCell(sheet[f"J{lastRow}"], punch_out.time(), normal_font, time_format, Border(), leftAlign)
                if punch[5]:
                    self.setCell(sheet[f"K{lastRow}"], "<- Unapproved Punch", normal_font, 'General', Border(), leftAlign)
                    anyUnapprovedPunches = True
                lastRow = lastRow + 1
                self.setCell(sheet[f"K{lastRow}"], "Elapsed Time:", bold_font, 'General', Border(), rightAlign)
                self.setCell(sheet[f"L{lastRow}"], shift_duration, normal_font, hrs_number_format, Border(), leftAlign)
                TotShopTime = TotShopTime + (shift_duration - lunch - TotConst - TotSer - TotOfc)
                TotalTime = TotalTime + shift_duration
                TotLunchTime = TotLunchTime + lunch
            #total all of the punches on the day
            self.setCell(sheet[f"H{TotalingRow}"], date.strftime('%A'), bold_font, 'General', Border(), Alignment())
            self.setCell(sheet[f"I{TotalingRow}"], date, normal_font, date_number_format, Border(), Alignment())
            self.setCell(sheet[f"J{TotalingRow}"], "Total:", bold_font, 'General', Border(), rightAlign)
            self.setCell(sheet[f"K{TotalingRow}"], TotalTime, normal_font, hrs_number_format, Border(), Alignment())
            self.setCell(sheet[f"L{TotalingRow}"], "Shop:", bold_font, 'General', Border(), rightAlign)
            self.setCell(sheet[f"M{TotalingRow}"], TotShopTime, normal_font, hrs_number_format, Border(), Alignment())
            self.setCell(sheet[f"N{TotalingRow}"], "Lunch:", bold_font, 'General', Border(), rightAlign)
            self.setCell(sheet[f"O{TotalingRow}"], TotLunchTime, normal_font, hrs_number_format, Border(), Alignment())
            self.setCell(sheet[f"P{TotalingRow}"], "Office:", bold_font, 'General', Border(), rightAlign)
            self.setCell(sheet[f"Q{TotalingRow}"], TotOfficeTime, normal_font, hrs_number_format, Border(), Alignment())
            self.setCell(sheet[f"R{TotalingRow}"], "Paid Hrs:", bold_font, 'General', Border(), rightAlign)
            self.setCell(sheet[f"S{TotalingRow}"], f"=K{TotalingRow}-O{TotalingRow}", normal_font, hrs_number_format, Border(), Alignment())
            lastRow = lastRow + 2
        sheet["C5"].value = TotalConstructionTime
        sheet["D5"].value = TotalServiceTime
        if anyUnapprovedPunches:
            self.setCell(sheet["A1"], "Notice there ARE unapproved Punches on this timecard", bold_font, 'General', Border(), leftAlign)
        return lastRow

    @discord.slash_command(name="timecardreport", description="Generate a report of all time punches for the previous week given an end date.")
    @commands.has_permissions(administrator=True)
    async def timecard_report(
        self,
        ctx: discord.ApplicationContext,
        week_end_date: discord.Option(str, description="End of Week Date in YYYY-MM-DD format [Must be a SATURDAY]"), # type: ignore
        employee_group: discord.Option(str, description="Employee group to include in the report", autocomplete=employee_group_autocomplete)  # type: ignore
    ):
        try:
            # Parse the provided date
            EOW = datetime.strptime(week_end_date, "%Y-%m-%d")
            week_start = EOW - timedelta(days=6)
            if not self.is_saturday(week_end_date):
                ctx.respond(f"{week_end_date} is not a saturday it is a: {self.get_day_of_week(week_end_date)}")
                return
            week_end = EOW + timedelta(days=1)

            # Connect to the database
            conn = sqlite3.connect(self.db)
            cursor = conn.cursor()

            # Fetch the group ID from the employee_group table
            cursor.execute("SELECT id FROM employee_group WHERE name = ?", (employee_group,))
            group_id = cursor.fetchone()
            if not group_id:
                await ctx.respond(f"Employee group '{employee_group}' not found.")
                return
            group_id = group_id[0]

            # Fetch the employee IDs for the selected group
            cursor.execute("SELECT employeeID FROM group_member WHERE groupID = ?", (group_id,))
            employee_ids = [row[0] for row in cursor.fetchall()]

            if not employee_ids:
                await ctx.respond(f"No employees found in group '{employee_group}'.")
                return

            # Prepare SQL query to fetch all punches for the previous week for the employees in the group
            placeholders = ",".join("?" for _ in employee_ids)
            sql_query = f"""
                SELECT 
                    e.id,
                    e.name,
                    pc.id,
                    pc.punchInTime,
                    pc.punchOutTime,
                    pc.punchInApproval,
                    pc.punchOutApproval,
                    pc.ignoreLunchBreak
                FROM punch_clock pc
                JOIN employee e ON pc.employeeID = e.id
                WHERE pc.punchInTime BETWEEN ? AND ? AND e.id IN ({placeholders})
                ORDER BY e.name, pc.punchInTime
            """

            params = [week_start, week_end] + employee_ids

            # Execute the query
            cursor.execute(sql_query, params)
            punches = cursor.fetchall()

            # Prepare to fetch work punches
            punch_data = {}
            employee_data = {}
            for punch in punches:
                employeeID, name, punch_id, punch_in, punch_out, in_approval, out_approval, ignore_lunch = punch
                punch = (name, punch_id, punch_in, punch_out, in_approval, out_approval, ignore_lunch)
                #Fetch work punches for each punch clock entry
                cursor.execute("""
                    SELECT
                        wt.punchType,
                        c.name,
                        wt.timeSpent
                    FROM work_time wt
                    JOIN customer c ON wt.customerID = c.id
                    WHERE wt.punchID = ?
                    ORDER BY wt.timeStarted
                """, (punch_id,))
                work_punches = cursor.fetchall()
                
                if name not in punch_data:
                    punch_data[name] = []
                    cursor.execute("""
                        SELECT
                            name,
                            addressLine1,
                            addressLine2,
                            addressCity,
                            addressState,
                            addressZip,
                            phoneNumber
                        FROM employee
                        WHERE id = ?
                    """, (employeeID,))
                    employee = cursor.fetchall()
                    employee_data[name] = employee[0]
                punch_data[name].append((punch, work_punches))

            conn.close()

            #if debugging print out the timeclock data
            if os.getenv('DEBUGGING'):
                print("DEBUGGING TIMECARD DATA:")
                for key, data in punch_data.items():
                    print(f"=-=-=-=-=-=-\n{key}")
                    for dat1, dat2 in data:
                        print(f"    {dat1}")
                        print(f"        {dat2}")
                print("End of DEBUGGING TIMECARD DATA")
                
            
            if not punches:
                print(f"No punches found for the week ending on {week_end_date}.")
                return

            # Create an Excel file
            file_path = f"reports/Weekly_Report_{week_end_date}.xlsx"
            self.createReportWorkbook(file_path, "Timecard")
            # Edit Excel File
            employees = [key for key in punch_data.keys()]
            # Create all the worksheets
            wb = load_workbook(file_path)
            ws = wb["Timecard"]
            ws.title = employees[0]
            # Create several copies of the template sheet within the new workbook
            for e in employees[1::]:
                # Copy the contentwb from the original new sheet to the new copy
                source = wb.active
                target = wb.copy_worksheet(source)
                target.title = e
            for e in employees:
                sheet = wb[e]
                sheet["D8"].value = week_end_date
                sheet["D11"].value = f"{employee_data[e][0]}"
                sheet["D14"].value = f"{employee_data[e][1]} {employee_data[e][2]}"
                sheet["D15"].value = f"{employee_data[e][3]}, {employee_data[e][4]} {employee_data[e][5]}"
                sheet["D18"].value = f"{employee_data[e][6]}"
                lastRow = self.reportTimecardData(sheet, punch_data[e])
                if lastRow < 1:
                    print(f"A critical error/bug occured in reportTimecardData on {e} because offset returned was {lastRow}!")
                else:
                    sheet.print_area = f"A1:G21 H1:S{lastRow}"
            wb.save(file_path)
            
            
            # Send the Excel file to the specified reports channel
            reports_channel_id = int(os.getenv('TIMECARD_REPORTS_CHANNEL_ID'))
            reports_channel = self.bot.get_channel(reports_channel_id)
            if reports_channel:
                await reports_channel.send(file=discord.File(file_path))
                await ctx.respond(f"Weekly report for the week ending on {week_end_date} has been generated and sent to the reports channel.", ephemeral=True)
                print(f"Weekly report for the week ending on {week_end_date} has been generated and sent to the reports channel.")
            else:
                print("Reports channel not found.")
        except ValueError as e:
            print(e)
            await ctx.respond("Invalid date format. Please use YYYY-MM-DD.")
        except Exception as e:
            print(e)
            await ctx.respond(f"An error occurred: {e}") 
        




    ## Database setup if it doesn't already exist
    def dbSetup(self, db):
        if not os.path.exists(db):
            f = open(db, "w")
            f.close()
            conn = sqlite3.connect(db)
            cursor = conn.cursor()
            #sets up employee type table
            cursor.execute('''
                CREATE TABLE employee_type (
	                id             INTEGER            PRIMARY KEY AUTOINCREMENT,
	                name           TEXT               NOT NULL,
	                rate           DECIMAL(10,5)      NOT NULL,
                    construction   BOOLEAN            NOT NULL DEFAULT "TRUE",
                    service        BOOLEAN            NOT NULL DEFAULT "TRUE", 
                    office         BOOLEAN            NOT NULL DEFAULT "FALSE"
                )
            ''')
            employeeTypesDefaultData = [
                ('Clerical', 1.5, False, False, True),
                ('Construction', 1.7, True, True, False),
                ('Salaried', 0.0, True, True, True)
            ]
            cursor.executemany('INSERT INTO employee_type (name, rate, construction, service, office) VALUES (?, ?, ?, ?, ?)', employeeTypesDefaultData)
            #sets up employee table
            cursor.execute('''
                CREATE TABLE employee (
                    id             UNSIGNED BIG INT    PRIMARY KEY,
                    name           TEXT                NOT NULL,
                    phoneNumber    TEXT                NOT NULL,
                    addressLine1   TEXT                NOT NULL,
                    addressLine2   TEXT                NOT NULL DEFAULT '',
                    addressCity    TEXT                NOT NULL,
                    addressState   TEXT                NOT NULL,
                    addressZip     TEXT                NOT NULL,
	                payrate        DECIMAL(10,2)       NOT NULL DEFAULT 16.00,
	                employeeTypeID INTEGER             NOT NULL DEFAULT 2,
	                lunchSkipable  BOOLEAN             NOT NULL DEFAULT "FALSE",
                    clockChannelId UNSIGNED BIG INT    NULL DEFAULT NULL,
                    clockMessageId UNSIGNED BIG INT    NULL DEFAULT NULL,
                    FOREIGN KEY (employeeTypeID) REFERENCES employee_type(id)
                )
            ''')
            #sets up punch clock table
            cursor.execute('''
                CREATE TABLE punch_clock (
                    id               UNSIGNED BIG INT    PRIMARY KEY,
                    employeeID       UNSIGNED BIG INT    NOT NULL,
                    punchInTime      DATETIME            NULL DEFAULT NULL,
                    punchInApproval  BOOLEAN             NOT NULL DEFAULT "TRUE",
                    punchOutTime     DATETIME            NULL DEFAULT NULL,
                    punchOutApproval BOOLEAN             NOT NULL DEFAULT "TRUE",
                    ignoreLunchBreak BOOLEAN             NOT NULL DEFAULT "FALSE",
                    checkChannelId   UNSIGNED BIG INT    NULL DEFAULT NULL,
                    checkMessageId   UNSIGNED BIG INT    NULL DEFAULT NULL,
                    FOREIGN KEY (employeeID) REFERENCES employee(id)
                )
            ''')
            #sets up the customer table
            cursor.execute('''
                CREATE TABLE customer (
                    id          INTEGER          PRIMARY KEY AUTOINCREMENT,
                    name        TEXT             NOT NULL
                )
            ''')
            #adds the default data for customer
            customerDefaultData = [
                (0, os.getenv('COMPANY_NAME'),)
            ]
            cursor.executemany('INSERT INTO customer (id, name) VALUES (?, ?)', customerDefaultData)
            # adds fake customers to the table for testing purposes
            if os.getenv('DEBUGGING'):
                customerDefaultData = [
                    ("Bond, James",),
                    ("Holmes, Sherlock",)
                ]
                cursor.executemany('INSERT INTO customer (name) VALUES (?)', customerDefaultData)
            #sets up work time table
            cursor.execute('''
                CREATE TABLE work_time (
                    id          UNSIGNED BIG INT                                                             PRIMARY KEY,
                    punchID     UNSIGNED BIG INT                                                             NOT NULL,
                    customerID  INTEGER                                                                      NOT NULL DEFAULT 0,
                    punchType   TEXT CHECK( punchType IN ('Construction','Service', 'Office') )              NOT NULL,
                    timeSpent   INTEGER CHECK( timeSpent >= 0 AND timeSpent <= 1440 AND timeSpent % 15 = 0)  NOT NULL DEFAULT 0,
                    timeStarted DATETIME                                                                     NOT NULL,
                    FOREIGN KEY (punchID) REFERENCES punch_clock(id),
                    FOREIGN KEY (customerID) REFERENCES customer(id)
                )
            ''')
            #sets up employee groups table
            cursor.execute('''
                CREATE TABLE employee_group (
                    id          INTEGER          PRIMARY KEY AUTOINCREMENT,
                    name        TEXT             NOT NULL
                )
            ''')
            #adds the default data for employee_group
            employeeGroupDefaultData = [
                (0, f"{str(os.getenv('COMPANY_NAME'))} Employee")
            ]
            cursor.executemany('INSERT INTO employee_group (id, name) VALUES (?, ?)', employeeGroupDefaultData)
            #sets up employee and employee groups LINK table
            cursor.execute('''
                CREATE TABLE group_member (
                    employeeID     UNSIGNED BIG INT    NOT NULL,
                    groupID        INTEGER             NOT NULL,
                    FOREIGN KEY (employeeID) REFERENCES employee(id),
                    FOREIGN KEY (groupID) REFERENCES employee_group(id)
                )
            ''')
            conn.commit()
            conn.close()


def setup(bot): # this is called by Pycord to setup the cog
    bot.add_cog(TimeTracking(bot)) # add the cog to the bot