"""Weekly Excel timecard generation.

Pure functions moved out of the cog. They do no DB or Discord I/O -- the cog
fetches data through the async db layer and hands it here, then runs
:func:`generate_timecard_report` in a worker thread (openpyxl/xlsxwriter are
blocking) so the event loop stays responsive.

Data shapes (unchanged from the original implementation):
    punch_data[name]   -> list of (punch, work_punches)
        punch          -> (name, punch_id, punch_in, punch_out,
                           in_approval, out_approval, ignore_lunch)
        work_punches   -> list of (punchType, customer_name, timeSpent_minutes)
    employee_data[name]-> (name, addr1, addr2, city, state, zip, phone)
"""

from __future__ import annotations

from copy import copy
from datetime import datetime, timedelta

import xlsxwriter as xwriter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

TEMPLATE_PATH = "templates/Template Sheets.xlsx"


# ---- date helpers (also used by slash-command autocomplete) ----------------

def round_to_quarter_hour(minutes: float) -> float:
    return round(minutes / 15) * 15


def convert_minutes_to_hours(minutes: float) -> float:
    return minutes / 60


def is_saturday(date_str: str) -> bool:
    return datetime.strptime(date_str, "%Y-%m-%d").weekday() == 5


def get_day_of_week(date_str: str) -> str:
    return datetime.strptime(date_str, "%Y-%m-%d").strftime("%A")


def get_closest_saturdays(date_object: datetime):
    days_until_saturday = (5 - date_object.weekday()) % 7
    closest_saturday = date_object + timedelta(days=days_until_saturday)
    saturdays = [closest_saturday + timedelta(weeks=i) for i in range(-14, 2)]
    saturdays = [s for s in saturdays if s <= closest_saturday + timedelta(weeks=1)]
    return [s.strftime("%Y-%m-%d") for s in saturdays]


def autofill_incomplete_date(user_input: str):
    today = datetime.today()
    if user_input == "":
        return today
    try:
        return datetime.strptime(user_input, "%Y-%m-%d")
    except ValueError:
        pass
    user_input = "".join(c for c in user_input if c.isnumeric() or c == "-")
    user_input = user_input if (not user_input or user_input[0] != "-") else user_input[1:]
    parts = user_input.split("-")
    if parts[-1] == "":
        if len(parts) == 1:
            return today
        if len(parts) == 2:
            user_input += f"{today.month:02d}-{today.day:02d}"
        elif len(parts) == 3:
            user_input += f"{today.day:02d}"
    else:
        if len(parts) > 3 or len(parts) < 1:
            return None
        if len(parts) == 1:
            if len(str(parts[0])) < len(str(today.year)):
                user_input += f"{str(today.year)[len(str(parts[0])):]}-{today.month:02d}-{today.day:02d}"
            else:
                user_input += f"-{today.month:02d}-{today.day:02d}"
        elif len(parts) == 2:
            mnth = f"{today.month:02d}"
            if len(str(parts[1])) < len(mnth):
                user_input += f"{mnth[len(str(parts[1])):]}-{today.day:02d}"
            else:
                user_input += f"-{mnth}-{today.day:02d}"
        else:
            dy = f"{today.day:02d}"
            if len(str(parts[2])) < len(dy):
                user_input += f"{dy[len(str(parts[2])):]}"
            else:
                user_input += f"-{dy}"
    try:
        return datetime.strptime(user_input, "%Y-%m-%d")
    except ValueError:
        return None


# ---- workbook helpers ------------------------------------------------------

def _set_cell(cell, value, font, number_format, border, alignment):
    cell.value = value
    cell.font = font
    cell.number_format = number_format
    cell.border = border
    cell.alignment = alignment


def _sort_key(element):
    punch, _work = element
    return datetime.strptime(punch[2], "%Y-%m-%d %H:%M:%S")


def _create_report_workbook(new_wb: str, template_sheet_name: str):
    workbook = xwriter.Workbook(new_wb)
    workbook.add_worksheet(template_sheet_name)
    workbook.close()
    tpwb = load_workbook(TEMPLATE_PATH)
    new_wb_obj = load_workbook(new_wb)
    tpws = tpwb[template_sheet_name]
    new_ws = new_wb_obj[template_sheet_name]
    for columns in tpws.iter_cols(min_row=1, min_col=1, max_col=tpws.max_column, max_row=1):
        for col in columns:
            letter = col.coordinate[:-1]
            new_ws.column_dimensions[letter].width = tpws.column_dimensions[letter].width
    for columns in tpws.iter_cols(min_row=1, min_col=1, max_col=tpws.max_column, max_row=tpws.max_row):
        for cell in columns:
            _set_cell(new_ws[cell.coordinate], cell.value, copy(cell.font),
                      copy(cell.number_format), copy(cell.border), copy(cell.alignment))
    new_wb_obj.save(new_wb)


def _report_timecard_data(sheet, data: list) -> int:
    last_row = 4
    data.sort(key=_sort_key)
    by_date = {}
    for tup in data:
        d = datetime.strptime(tup[0][2], "%Y-%m-%d %H:%M:%S").date()
        by_date.setdefault(d, []).append(tup)

    any_unapproved = False
    normal_font, bold_font = Font(name="Arial", size=10), Font(name="Arial", size=10, bold=True)
    date_fmt, hrs_fmt, time_fmt = "mm-dd-yy", '0.00 "hrs"', "hh:mm AM/PM"
    wt_border = Border(top=Side(border_style="thin", color="00000000"),
                       bottom=Side(border_style="thin", color="00000000"))
    right, left = Alignment(horizontal="right"), Alignment(horizontal="left")

    for date in by_date:
        totaling_row = last_row
        total_time = tot_office = tot_shop = tot_lunch = 0
        for punch, work in by_date[date]:
            punch_in = datetime.strptime(punch[2], "%Y-%m-%d %H:%M:%S")
            punch_out = datetime.strptime(punch[3], "%Y-%m-%d %H:%M:%S") if punch[3] else datetime.now()
            shift = convert_minutes_to_hours(round_to_quarter_hour((punch_out - punch_in).total_seconds() / 60))
            lunch = 0.5 if not punch[6] and shift >= 6.0 else 0
            construction, service = [], []
            tot_c = tot_s = tot_o = 0
            for worktype, cust, minutes in work or []:
                hrs = minutes / 60
                if worktype == "Office":
                    tot_o += hrs
                    tot_office += hrs
                elif worktype == "Service":
                    tot_s += hrs
                    service.append((cust, hrs))
                else:
                    tot_c += hrs
                    construction.append((cust, hrs))

            last_row += 1
            _set_cell(sheet[f"I{last_row}"], "Clock-In", bold_font, "General", Border(), Alignment())
            _set_cell(sheet[f"J{last_row}"], punch_in.time(), normal_font, time_fmt, Border(), left)
            if not bool(punch[4]):
                _set_cell(sheet[f"K{last_row}"], "<- Unapproved Punch", normal_font, "General", Border(), left)
                any_unapproved = True
            if construction:
                last_row += 1
                _set_cell(sheet[f"J{last_row}"], "Construction", bold_font, "General", wt_border, right)
                _set_cell(sheet[f"K{last_row}"], tot_c, normal_font, hrs_fmt, wt_border, Alignment())
                for name, hrs in construction:
                    last_row += 1
                    _set_cell(sheet[f"J{last_row}"], name, normal_font, "General", Border(), right)
                    _set_cell(sheet[f"K{last_row}"], hrs, normal_font, hrs_fmt, Border(), Alignment())
            if service:
                last_row += 1
                _set_cell(sheet[f"J{last_row}"], "Service", bold_font, "General", wt_border, right)
                _set_cell(sheet[f"K{last_row}"], tot_s, normal_font, hrs_fmt, wt_border, Alignment())
                for name, hrs in service:
                    last_row += 1
                    _set_cell(sheet[f"J{last_row}"], name, normal_font, "General", Border(), right)
                    _set_cell(sheet[f"K{last_row}"], hrs, normal_font, hrs_fmt, Border(), Alignment())
            last_row += 1
            _set_cell(sheet[f"I{last_row}"], "Clock-Out", bold_font, "General", Border(), Alignment())
            _set_cell(sheet[f"J{last_row}"], punch_out.time(), normal_font, time_fmt, Border(), left)
            if not bool(punch[5]):
                _set_cell(sheet[f"K{last_row}"], "<- Unapproved Punch", normal_font, "General", Border(), left)
                any_unapproved = True
            last_row += 1
            _set_cell(sheet[f"K{last_row}"], "Elapsed Time:", bold_font, "General", Border(), right)
            _set_cell(sheet[f"L{last_row}"], shift, normal_font, hrs_fmt, Border(), left)
            tot_shop += shift - lunch - tot_c - tot_s - tot_o
            total_time += shift
            tot_lunch += lunch

        _set_cell(sheet[f"H{totaling_row}"], date.strftime("%A"), bold_font, "General", Border(), Alignment())
        _set_cell(sheet[f"I{totaling_row}"], date, normal_font, date_fmt, Border(), Alignment())
        _set_cell(sheet[f"J{totaling_row}"], "Total:", bold_font, "General", Border(), right)
        _set_cell(sheet[f"K{totaling_row}"], total_time, normal_font, hrs_fmt, Border(), Alignment())
        _set_cell(sheet[f"L{totaling_row}"], "Shop:", bold_font, "General", Border(), right)
        _set_cell(sheet[f"M{totaling_row}"], tot_shop, normal_font, hrs_fmt, Border(), Alignment())
        _set_cell(sheet[f"N{totaling_row}"], "Lunch:", bold_font, "General", Border(), right)
        _set_cell(sheet[f"O{totaling_row}"], tot_lunch, normal_font, hrs_fmt, Border(), Alignment())
        _set_cell(sheet[f"M{totaling_row + 1}"], "Office:", bold_font, "General", Border(), right)
        _set_cell(sheet[f"N{totaling_row + 1}"], tot_office, normal_font, hrs_fmt, Border(), Alignment())
        _set_cell(sheet[f"O{totaling_row + 1}"], "Paid Hrs:", bold_font, "General", Border(), right)
        _set_cell(sheet[f"P{totaling_row + 1}"], f"=K{totaling_row}-O{totaling_row}", normal_font, hrs_fmt, Border(), Alignment())
        last_row += 2

    if any_unapproved:
        _set_cell(sheet["A1"], "Notice there ARE unapproved Punches on this timecard", bold_font, "General", Border(), left)
    return last_row


def generate_timecard_report(file_path: str, employees: list, punch_data: dict,
                             employee_data: dict, week_end_date: str):
    """Build the multi-sheet weekly workbook. Blocking; run in a thread."""
    _create_report_workbook(file_path, "Timecard")
    wb = load_workbook(file_path)
    ws = wb["Timecard"]
    ws.title = employees[0]
    for e in employees[1:]:
        target = wb.copy_worksheet(wb.active)
        target.title = e
    for e in employees:
        sheet = wb[e]
        info = employee_data[e]
        sheet["D8"].value = week_end_date
        sheet["D11"].value = f"{info[0]}"
        sheet["D14"].value = f"{info[1]} {info[2]}"
        sheet["D15"].value = f"{info[3]}, {info[4]} {info[5]}"
        sheet["D18"].value = f"{info[6]}"
        last_row = _report_timecard_data(sheet, punch_data[e])
        if last_row >= 1:
            sheet.print_area = f"A1:G21 H1:P{last_row}"
            sheet.page_margins.left = sheet.page_margins.right = 0.25
            sheet.page_margins.top = sheet.page_margins.bottom = 0.75
            sheet.page_margins.header = sheet.page_margins.footer = 0.3
            sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            sheet.page_setup.fitToHeight = False
    wb.save(file_path)
