"""Hipp staffing invoice workbook.

Two sheets, mirroring the owner's template:

* **Hipp Invoice** — one row per employee: current Pay Rate, Std/OT hours and
  billed rates (payrate x employee_type markup, OT x1.5) and the line total, with
  a grand total.
* **Payroll by Department** — each employee's total pay distributed across the work
  categories by hours (Pool = the bot's Construction, plus Service / Office), with
  **Shop** absorbing the rounding remainder.

Layout/styling: content starts at B2 (an empty column A / row 1 give breathing room
from the edge). Time and money data are centered. Zebra data rows (white / "White,
Background 2, Darker 5%" = #DCDADA), a "Darker 15%" (#C5C3C3) total row, and
zero-valued data cells shown in #C5C3C3 so they recede. Borders: the Invoice table
is enclosed with a vertical line on every column plus a boxed header and boxed total
band; Payroll by Department keeps vertical lines only on the key columns (Emp #,
Employee, Total Hrs, Total Pay) with the whole table, header row, and total row each
outer-bordered. Each sheet sets fit-to-width printing and repeats the header rows as
print titles. Gray hexes are the modern Office theme's Background-2 tints, and the
modern Office theme is embedded so the in-Excel palette matches the owner's sheets
(openpyxl bundles the older theme).

Built programmatically (openpyxl) so no real accounting data lives in the repo.
Blocking; call via asyncio.to_thread. Rows come from costing.hipp_billing.
"""

from __future__ import annotations

import re
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

from . import costing

# Offset so the top-left of every table sits at B2 (empty col A + row 1 for spacing).
_COL_SHIFT = 1
_ROW_SHIFT = 1

_HDR = Font(name="Arial", size=10, bold=True)
_BODY = Font(name="Arial", size=10)
_TITLE = Font(name="Arial", size=14, bold=True)
_ZERO = Font(name="Arial", size=10, color="C5C3C3")       # muted zeros
_CUR = '$#,##0.00'
_HRS = '0.00'
_HRS_TOTAL = '0.00 "hrs";[Red]-0.00 "hrs";"-"'             # Total-Hrs column format
_thin = Side(style="thin")
_HFILL = PatternFill("solid", fgColor="D9E1F2")            # header
_WHITE = PatternFill("solid", fgColor="FFFFFF")            # zebra: white row
_ALT = PatternFill("solid", fgColor="DCDADA")             # White, Background 2, Darker 5%
_TOTAL_FILL = PatternFill("solid", fgColor="C5C3C3")       # White, Background 2, Darker 15%
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_MID = Alignment(horizontal="center", vertical="center")   # centered data (no wrap)
_RIGHT = Alignment(horizontal="right")


def _shift_col(letter: str) -> str:
    return get_column_letter(column_index_from_string(letter) + _COL_SHIFT)


def _addr(col: str, row: int) -> str:
    return f"{_shift_col(col)}{row + _ROW_SHIFT}"


def _modern_theme() -> bytes | None:
    """openpyxl bundles the OLD Office 2007-2010 theme, so a downloaded report shows
    that palette (beige "Background 2", muted accents) instead of the modern one the
    owner's spreadsheets use. Rebuild the theme with the modern Office color scheme
    (+ Calibri Light headings) so the in-Excel palette matches. Colors in the report
    itself are explicit hex, so this only affects the theme palette, never the look.
    Returns theme bytes to assign to Workbook.loaded_theme, or None if unavailable."""
    try:
        from openpyxl.writer.theme import theme_xml
    except Exception:  # noqa: BLE001 - never let theming break report generation
        return None
    xml = theme_xml.decode() if isinstance(theme_xml, (bytes, bytearray)) else theme_xml
    modern_clr = (
        '<a:clrScheme name="Office">'
        '<a:dk1><a:sysClr val="windowText" lastClr="000000"/></a:dk1>'
        '<a:lt1><a:sysClr val="window" lastClr="FFFFFF"/></a:lt1>'
        '<a:dk2><a:srgbClr val="44546A"/></a:dk2>'
        '<a:lt2><a:srgbClr val="E7E6E6"/></a:lt2>'
        '<a:accent1><a:srgbClr val="4472C4"/></a:accent1>'
        '<a:accent2><a:srgbClr val="ED7D31"/></a:accent2>'
        '<a:accent3><a:srgbClr val="A5A5A5"/></a:accent3>'
        '<a:accent4><a:srgbClr val="FFC000"/></a:accent4>'
        '<a:accent5><a:srgbClr val="5B9BD5"/></a:accent5>'
        '<a:accent6><a:srgbClr val="70AD47"/></a:accent6>'
        '<a:hlink><a:srgbClr val="0563C1"/></a:hlink>'
        '<a:folHlink><a:srgbClr val="954F72"/></a:folHlink>'
        '</a:clrScheme>'
    )
    xml = re.sub(r"<a:clrScheme.*?</a:clrScheme>", lambda _m: modern_clr, xml, count=1, flags=re.S)
    xml = xml.replace('<a:latin typeface="Cambria"/>', '<a:latin typeface="Calibri Light"/>', 1)
    return xml.encode("utf-8")


_MODERN_THEME = _modern_theme()


def _border(sides) -> Border:
    """Border with only the named sides ('left'/'right'/'top'/'bottom') drawn thin."""
    return Border(**{s: _thin for s in sides})


def _cell(ws, col, row, value, font=_BODY, fmt="General", align=None):
    """Plain cell (letterhead) — no fill, no border. Logical (col, row); offset applied."""
    c = ws[_addr(col, row)]
    c.value = value
    c.font = font
    c.number_format = fmt
    if align:
        c.alignment = align
    return c


def _put(ws, col, r, value, fmt, align, fill, sides, base_font=_BODY, zero_gray=False):
    c = ws[_addr(col, r)]
    c.value = value
    c.font = _ZERO if (zero_gray and isinstance(value, (int, float)) and value == 0) else base_font
    c.number_format = fmt
    c.fill = fill
    c.border = _border(sides)
    if align:
        c.alignment = align
    return c


def _finalize(ws, header_last_row, last_col, last_row, landscape):
    """Set the print area to the used range ($B$1 -> bottom-right edited cell), fit
    to width, modest margins, and repeat the top spacing + header rows on every page.
    `last_col` is the logical rightmost column; `last_row` is the actual bottom row."""
    ws.print_area = f"${_shift_col('A')}$1:${_shift_col(last_col)}${last_row}"
    ws.print_title_rows = f"1:{header_last_row + _ROW_SHIFT}"
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5, header=0.3, footer=0.3)
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def _build_invoice_sheet(ws, rows, period_label, invoice_number):
    ws.title = "Hipp Invoice"
    _cell(ws, "A", 1, "HIPP Temporary Skills, Inc.", _TITLE)
    _cell(ws, "A", 2, "Payroll billing — SwimShack Inc.", _BODY)
    _cell(ws, "A", 3, f"Invoice #: {invoice_number or ''}", _BODY)
    _cell(ws, "E", 3, f"Date: {date.today().isoformat()}", _BODY)
    _cell(ws, "A", 4, f"Period Covered: {period_label}", _BODY)

    first, last = "A", "H"
    top = 6
    headers = [("A", "Emp. #"), ("B", "Employee"), ("C", "Pay Rate"), ("D", "Std. Hrs."),
               ("E", "Std. Rate"), ("F", "OT Hrs."), ("G", "OT Rate"), ("H", "Total")]
    widths = [8, 26, 11, 10, 12, 10, 12, 14]
    for (col, text), w in zip(headers, widths):
        _put(ws, col, top, text, "General", _CENTER, _HFILL, {"left", "right", "top", "bottom"}, _HDR)
        ws.column_dimensions[_shift_col(col)].width = w

    r = top
    grand = 0.0
    for n, row in enumerate(rows, start=1):
        r += 1
        fill = _WHITE if n % 2 == 1 else _ALT
        cells = [("A", n, "General", _MID), ("B", row["name"], "General", None),
                 ("C", row.get("base_rate", 0.0), _CUR, _MID), ("D", row["std_hrs"], _HRS, _MID),
                 ("E", row["std_rate"], _CUR, _MID), ("F", row["ot_hrs"], _HRS, _MID),
                 ("G", row["ot_rate"], _CUR, _MID), ("H", row["total"], _CUR, _MID)]
        for col, val, fmt, align in cells:
            _put(ws, col, r, val, fmt, align, fill, {"left", "right"}, zero_gray=True)
        grand += row["total"]
    r += 1
    total_cells = [("A", None, "General", None), ("B", "Total", "General", _RIGHT),
                   ("C", None, _CUR, None), ("D", None, _HRS, None), ("E", None, _CUR, None),
                   ("F", None, _HRS, None), ("G", None, _CUR, None), ("H", round(grand, 2), _CUR, _MID)]
    for col, val, fmt, align in total_cells:
        sides = {"top", "bottom"} | ({"left"} if col == first else set()) | ({"right"} if col == last else set())
        _put(ws, col, r, val, fmt, align, _TOTAL_FILL, sides, _HDR)
    _finalize(ws, header_last_row=top, last_col="H", last_row=r + _ROW_SHIFT, landscape=False)


def _build_department_sheet(ws, rows):
    ws.title = "Payroll by Department"
    vcols = {"A", "B", "G", "L"}   # Emp #, Employee, Total Hrs, Total Pay
    top = 1
    headers = [("A", "Emp. #"), ("B", "Employee"),
               ("C", "Shop\nHrs"), ("D", "Pool\nHrs"), ("E", "Service\nHrs"), ("F", "Office\nHrs"), ("G", "Total\nHrs"),
               ("H", "Shop $"), ("I", "Pool $"), ("J", "Service $"), ("K", "Office $"), ("L", "Total Pay")]
    widths = [8, 24, 8, 8, 9, 8, 9, 12, 12, 12, 12, 14]
    for (col, text), w in zip(headers, widths):
        sides = {"top", "bottom"} | ({"left", "right"} if col in vcols else set())
        _put(ws, col, top, text, "General", _CENTER, _HFILL, sides, _HDR)
        ws.column_dimensions[_shift_col(col)].width = w

    r = top
    tot = {k: 0.0 for k in ("shop_h", "pool_h", "svc_h", "off_h", "th",
                            "shop_d", "pool_d", "svc_d", "off_d", "tp")}
    for n, row in enumerate(rows, start=1):
        ch = row["category_hours"]
        shop_h, pool_h = ch[costing.SHOP], ch["Construction"]
        svc_h, off_h = ch["Service"], ch["Office"]
        total_h = shop_h + pool_h + svc_h + off_h
        dist = costing.distribute_pay(
            row["total"], {"Pool": pool_h, "Service": svc_h, "Office": off_h, "Shop": shop_h}, "Shop")
        r += 1
        fill = _WHITE if n % 2 == 1 else _ALT
        cells = [("A", n, "General", _MID), ("B", row["name"], "General", None),
                 ("C", shop_h, _HRS, _MID), ("D", pool_h, _HRS, _MID), ("E", svc_h, _HRS, _MID),
                 ("F", off_h, _HRS, _MID), ("G", round(total_h, 2), _HRS_TOTAL, _MID),
                 ("H", dist["Shop"], _CUR, _MID), ("I", dist["Pool"], _CUR, _MID),
                 ("J", dist["Service"], _CUR, _MID), ("K", dist["Office"], _CUR, _MID),
                 ("L", row["total"], _CUR, _MID)]
        for col, val, fmt, align in cells:
            sides = {"left", "right"} if col in vcols else set()
            _put(ws, col, r, val, fmt, align, fill, sides, zero_gray=True)
        tot["shop_h"] += shop_h; tot["pool_h"] += pool_h; tot["svc_h"] += svc_h
        tot["off_h"] += off_h; tot["th"] += total_h
        tot["shop_d"] += dist["Shop"]; tot["pool_d"] += dist["Pool"]
        tot["svc_d"] += dist["Service"]; tot["off_d"] += dist["Office"]; tot["tp"] += row["total"]
    r += 1
    total_cells = [("A", None, "General", None), ("B", "Totals", "General", _RIGHT),
                   ("C", round(tot["shop_h"], 2), _HRS, _MID), ("D", round(tot["pool_h"], 2), _HRS, _MID),
                   ("E", round(tot["svc_h"], 2), _HRS, _MID), ("F", round(tot["off_h"], 2), _HRS, _MID),
                   ("G", round(tot["th"], 2), _HRS_TOTAL, _MID), ("H", round(tot["shop_d"], 2), _CUR, _MID),
                   ("I", round(tot["pool_d"], 2), _CUR, _MID), ("J", round(tot["svc_d"], 2), _CUR, _MID),
                   ("K", round(tot["off_d"], 2), _CUR, _MID), ("L", round(tot["tp"], 2), _CUR, _MID)]
    for col, val, fmt, align in total_cells:
        sides = {"top", "bottom"} | ({"left", "right"} if col in vcols else set())
        _put(ws, col, r, val, fmt, align, _TOTAL_FILL, sides, _HDR)
    _finalize(ws, header_last_row=top, last_col="L", last_row=r + _ROW_SHIFT, landscape=True)


def generate_hipp_invoice(file_path: str, rows: list[dict], period_label: str,
                          invoice_number: str | None = None) -> None:
    """Write the two-sheet Hipp invoice workbook to ``file_path``. ``rows`` is a list
    of costing.hipp_billing dicts. Blocking — run in a worker thread."""
    wb = Workbook()
    if _MODERN_THEME:
        wb.loaded_theme = _MODERN_THEME   # match the owner's spreadsheets' theme palette
    _build_invoice_sheet(wb.active, rows, period_label, invoice_number)
    _build_department_sheet(wb.create_sheet("Payroll by Department"), rows)
    wb.save(file_path)
